"""Utility to compare OCR/LLM JSON outputs against the ground-truth Excel sheet.

Usage:
    python tools/benchmark_results.py --standard standart.xlsx --uploads uploads
"""
from __future__ import annotations

import argparse
import json
import re
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

from openpyxl import load_workbook

TARGET_FIELDS = [
    "Название_файла",
    "Тип_документа",
    "Номер_документа",
    "Дата_документа",
    "Наименование_заказчика",
    "Наименование_исполнителя",
    "ИНН_заказчика",
    "ИНН_исполнителя",
    "КПП_заказчика",
    "КПП_исполнителя",
    "Адрес_заказчика",
    "Адрес_исполнителя",
]

STRICT_NUMERIC_FIELDS = {
    "ИНН_заказчика",
    "ИНН_исполнителя",
    "КПП_заказчика",
    "КПП_исполнителя",
    "Номер_документа",
}

WHITESPACE_ONLY_FIELDS = {
    "Адрес_заказчика",
    "Адрес_исполнителя",
    "Наименование_заказчика",
    "Наименование_исполнителя",
}


def normalize_value(value: Optional[str], field: str) -> str:
    if value is None:
        return ""

    text = str(value).strip()
    text = text.replace("\u00a0", " ")
    text = re.sub(r"\s+", " ", text)
    text = text.replace("Ё", "Е").replace("ё", "е")

    if field in STRICT_NUMERIC_FIELDS:
        digits = re.sub(r"\D", "", text)
        return digits

    if field in WHITESPACE_ONLY_FIELDS:
        return text.lower()

    return text.lower()


def build_expected_rows(standard_path: Path) -> List[Dict[str, str]]:
    wb = load_workbook(standard_path)
    ws = wb.active
    header_row = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    normalized_headers = [str(h).strip() if h is not None else "" for h in header_row]

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(value is None for value in row):
            continue
        row_dict = {
            header: value
            for header, value in zip(normalized_headers, row)
            if header in TARGET_FIELDS
        }
        if row_dict:
            rows.append(row_dict)
    return rows


def guess_upload_name(filename: str) -> str:
    stem = Path(filename).stem.lstrip("_")
    digits = re.sub(r"\D", "", stem)
    if digits:
        index = int(digits)
        return f"{index:03d}_result.json"
    return f"{stem}_result.json"


@dataclass
class ComparisonResult:
    matched: int
    total: int
    missing_files: List[str]
    field_mismatches: Dict[str, List[Tuple[str, str, str]]]

    @property
    def accuracy(self) -> float:
        if self.total == 0:
            return 0.0
        return self.matched / self.total


def compare_results(expected_rows: Iterable[Dict[str, str]], uploads_dir: Path) -> ComparisonResult:
    matched = 0
    total = 0
    missing_files: List[str] = []
    field_mismatches: Dict[str, List[Tuple[str, str, str]]] = defaultdict(list)

    for expected in expected_rows:
        filename = str(expected.get("Название_файла", ""))
        json_name = guess_upload_name(filename)
        json_path = uploads_dir / json_name

        if not json_path.exists():
            missing_files.append(json_name)
            total += len(TARGET_FIELDS)
            continue

        try:
            payload = json.loads(json_path.read_text(encoding="utf-8"))
        except json.JSONDecodeError as exc:
            missing_files.append(f"{json_name} (invalid JSON: {exc})")
            total += len(TARGET_FIELDS)
            continue

        for field in TARGET_FIELDS:
            expected_value = normalize_value(expected.get(field), field)
            actual_value = normalize_value(payload.get(field), field)
            total += 1
            if expected_value == actual_value:
                matched += 1
            else:
                field_mismatches[field].append((filename, expected_value, actual_value))

    return ComparisonResult(matched, total, missing_files, field_mismatches)


def print_report(result: ComparisonResult) -> None:
    accuracy_pct = result.accuracy * 100
    print(f"Общая точность: {accuracy_pct:.2f}% ({result.matched}/{result.total})")

    if result.missing_files:
        print("\nФайлы без валидного JSON:")
        for name in result.missing_files:
            print(f"  - {name}")

    if not result.field_mismatches:
        print("\nВсе поля соответствуют эталону.")
        return

    print("\nТоп расхождений по полям:")
    for field, mismatches in result.field_mismatches.items():
        print(f"\n{field} (несовпадений: {len(mismatches)})")
        for filename, expected_value, actual_value in mismatches[:5]:
            print(f"  {filename}: ожидалось '{expected_value}' | получено '{actual_value}'")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Compare OCR JSON outputs with the standard Excel file.")
    parser.add_argument("--standard", type=Path, default=Path("standart.xlsx"), help="Путь к эталонной таблице (Excel)")
    parser.add_argument("--uploads", type=Path, default=Path("uploads"), help="Папка с результатами JSON")
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    if not args.standard.exists():
        raise SystemExit(f"Эталонный файл не найден: {args.standard}")
    if not args.uploads.exists():
        raise SystemExit(f"Папка с результатами не найдена: {args.uploads}")

    expected_rows = build_expected_rows(args.standard)
    comparison = compare_results(expected_rows, args.uploads)
    print_report(comparison)


if __name__ == "__main__":
    main()
